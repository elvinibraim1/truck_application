from flask import Flask, render_template, request, send_file
import os
from PIL import Image
import pytesseract
from openpyxl import Workbook
import uuid
import re

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        files = request.files.getlist('images')
        data_rows = []

        for file in files:
            filepath = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(filepath)

            image = Image.open(filepath)
            text = pytesseract.image_to_string(image)

            data = extract_data_from_text(text)
            data_rows.append(data)

        filename = f"{uuid.uuid4().hex}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, filename)
        save_to_excel(data_rows, excel_path)

        return send_file(excel_path, as_attachment=True)

    return render_template('index.html')

def extract_data_from_text(text):
    data = ""
    brut = ""
    tara = ""
    net = ""
    marfa = ""
    nr_tractor = ""
    nr_remorca = ""
    ora_intrare = ""
    ora_iesire = ""
    sofer = ""
    furnizor = ""
    beneficiar = ""

    lines = text.split('\n')
    for line in lines:
        line = line.strip()

        if "MARFA" in line:
            marfa = line.split(":")[-1].strip()

        elif "NR. TRACTOR" in line:
            match = re.search(r'TRACTOR:\s*([A-Z0-9\-]+)', line)
            if match:
                nr_tractor = match.group(1)

        elif "NR. REMORCA" in line:
            match = re.search(r'REMORCA:\s*([A-Z0-9\-]+)', line)
            if match:
                nr_remorca = match.group(1)

        elif "BRUT" in line:
            match = re.search(r'BRUT\s*[:\-]?\s*([\d.]+)', line)
            if match:
                brut = match.group(1)

        elif "TARA" in line:
            match = re.search(r'TARA\s*[:\-]?\s*([\d.]+)', line)
            if match:
                tara = match.group(1)

        elif "NET" in line:
            match = re.search(r'NET\s*[:\-]?\s*([\d.]+)', line)
            if match:
                net = match.group(1)

        elif re.search(r'\b\d{2}\.\d{2}\.\d{4}\b', line):
            match = re.search(r'(\d{2}\.\d{2}\.\d{4})', line)
            if match:
                data = match.group(1)

        elif "ORA:" in line or "ORA" in line:
            times = re.findall(r'\d{2}:\d{2}:\d{2}', line)
            if times:
                if not ora_intrare:
                    ora_intrare = times[0]
                elif not ora_iesire and len(times) > 1:
                    ora_iesire = times[1]

        elif "SOFER" in line:
            sofer = line.split(":")[-1].strip()

        elif "FURNIZOR" in line:
            furnizor = line.split(":")[-1].strip()

        elif "BENEFICIAR" in line:
            beneficiar = line.split(":")[-1].strip()

    return [
        data,
        brut,
        tara,
        net,
        marfa,
        nr_tractor,
        nr_remorca,
        ora_intrare,
        ora_iesire,
        sofer,
        furnizor,
        beneficiar
    ]


def save_to_excel(rows, path):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    headers = [
        "Data",
        "Brut (KG)",
        "Tara (KG)",
        "Net (KG)",
        "Marfa",
        "Nr. Tractor", "Nr. Remorca",
        "Ora Intrare", "Ora Iesire",
        "Sofer", "Furnizor", "Beneficiar"
    ]
    ws.append(headers)

    # Adaugă datele
    for row in rows:
        ws.append(row)

    # Stilizare header
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    alignment = Alignment(horizontal="center", vertical="center")

    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for line in lines:
        line = line.strip()

        if "MARFA" in line:
            marfa = line.split(":")[-1].strip()

        elif "NR. TRACTOR" in line:
            match = re.search(r'TRACTOR:\s*([A-Z0-9\-]+)', line)
            if match:
                nr_tractor = match.group(1)

        elif "NR. REMORCA" in line:
            match = re.search(r'REMORCA:\s*([A-Z0-9\-]+)', line)
            if match:
                nr_remorca = match.group(1)

        elif "BRUT" in line:
            match = re.search(r'BRUT\s*[:\-]?\s*([\d.]+)', line)
            if match:
                brut = match.group(1)

        elif "TARA" in line:
            match = re.search(r'TARA\s*[:\-]?\s*([\d.]+)', line)
            if match:
                tara = match.group(1)

        elif "NET" in line:
            match = re.search(r'NET\s*[:\-]?\s*([\d.]+)', line)
            if match:
                net = match.group(1)

        elif re.search(r'\b\d{2}\.\d{2}\.\d{4}\b', line):
            match = re.search(r'(\d{2}\.\d{2}\.\d{4})', line)
            if match:
                data = match.group(1)

        elif "ORA:" in line or "ORA" in line:
            times = re.findall(r'\d{2}:\d{2}:\d{2}', line)
            # Asigură-te că ambele coloane există mereu
            if times:
                ora_intrare = times[0]
                if len(times) > 1:
                    ora_iesire = times[1]
                else:
                    ora_iesire = ""

        elif "SOFER" in line:
            sofer = line.split(":")[-1].strip()

        elif "FURNIZOR" in line:
            furnizor = line.split(":")[-1].strip()

        elif "BENEFICIAR" in line:
            beneficiar = line.split(":")[-1].strip()
